from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from scaffold_job import ScaffoldJob
from job_dict import JobDictionary
from datetime import datetime


class ExcelHandler:
    def __init__(self, file_name: str, jobs: JobDictionary):
        self.file_name = file_name
        self.jobs = jobs

    def load_from_excel(self):
        try:
            workbook = load_workbook(self.file_name) #Loads the excel file as workbook
            sheet = workbook.active #Changes name of excel file to sheet

            # Get job names from the JobNames sheet
            job_sheet = workbook['JobNames'] 
            job_names = [row[0] for row in job_sheet.iter_rows(values_only=True) if row[0] is not None] #Stores the names of the jobs from the jobsheet in a list

            current_job_name = None #Creates a variable to store the current job name 
            date_row = [cell.value for cell in sheet[2]] #Creates a list of all the dates found in the second row

            for row in sheet.iter_rows(min_row=1, values_only=True):  # Iterates through the rows in the sheet, starting from the first
                first_cell = row[0] #The first cell in the current row is stored as first_cell (either a jobname or scaffold item)

                # Check if the first cell is a job name
                if first_cell and first_cell.strip() in job_names:
                    current_job_name = first_cell.strip() #If the first cell in this row is a job name, it sets it to the current job name
                    if current_job_name not in self.jobs.jobs: #If the job name isn't already stored in the job dict, it creates a new job name
                        self.jobs.jobs[current_job_name] = ScaffoldJob(current_job_name)

                # If not a job name, it should be a scaffold item
                elif current_job_name:
                    item = row[0]
                    if item: #If the cell isn't empty
                        item = item.strip() #Removes any whitespace before or after the item
                        current_job = self.jobs.jobs[current_job_name] #Sets the current job to the job name found in the job dict
                        for col, quantity in enumerate(row[1:], start=1): #Iterates over each cell in the row
                            if quantity is not None:
                                date_str = date_row[col] #Matches the quantity added to the date it was added, stores the date as a string
                                if date_str: #If the date is not empty
                                    try:
                                        date_added = datetime.strptime(date_str, '%d/%m/%Y').date() #Converts the date string to a date object
                                        if quantity >= 0:
                                            current_job.scaffold[item] = current_job.scaffold.get(item, 0) + quantity #Adds the quantity to the total quantity stored of this item
                                            current_job.dates[item] = date_added #Stores the date that this item was added
                                        else:
                                            if item in current_job.deletions: #If the quantity is negative and the item is in the deletions dict
                                                current_job.deletions[item].append((-quantity, date_added)) #Adds the amount deleted and the date of deletion
                                            else:
                                                current_job.deletions[item] = [(-quantity, date_added)] #If the item is not currently in the deletions dict, adds it
                                    except ValueError as ve:
                                        print(f"Date parsing error for {date_str}: {ve}")
            print(f"Data successfully loaded from {self.file_name}")
        except FileNotFoundError:
            print(f"{self.file_name} not found. Starting with an empty dataset.")
        except Exception as e:
            print(f"An error occurred while loading data from {self.file_name}: {e}")
#Data is loaded and stored in scaffoldjob items

    def save_to_excel(self):
        try:
            workbook = load_workbook(self.file_name)
        except FileNotFoundError:
            workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Scaffold Data"

        if 'JobNames' in workbook.sheetnames:
            job_sheet = workbook['JobNames']
        else:
            job_sheet = workbook.create_sheet('JobNames')

        job_names = list(self.jobs.jobs.keys())
        dates = set()
        for job in self.jobs.jobs.values():
            dates.update(job.dates.values())
            for item_deletions in job.deletions.values():
                dates.update(date for _, date in item_deletions)
        dates = sorted(dates)

        # Write job names to JobNames sheet
        job_sheet.delete_rows(1, job_sheet.max_row)
        for row, job_name in enumerate(job_names, start=1):
            job_sheet.cell(row=row, column=1, value=job_name)

        # Initialize row number
        row_num = 1

        for job_name in job_names:
            job = self.jobs.jobs[job_name]
            # Write job name in a merged cell
            sheet.cell(row=row_num, column=1, value=job_name)
            if len(dates) > 1:
                sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(dates) + 1)
            
            row_num += 1

            # Write dates in the second row, with new dates added as new columns
            existing_dates = [cell.value for cell in sheet[2] if cell.value]
            new_dates = [date.strftime('%d/%m/%Y') for date in dates if date.strftime('%d/%m/%Y') not in existing_dates]

            date_columns = {}
            for col, date_str in enumerate(existing_dates + new_dates, start=2):
                sheet.cell(row=2, column=col, value=date_str)
                date_columns[date_str] = col
            
            row_num += 1

            # Write scaffold items and quantities
            items_written = {}
            for item in job.scaffold:
                if item not in items_written:
                    sheet.cell(row=row_num, column=1, value=item)
                    items_written[item] = row_num
                    row_num += 1
                row_index = items_written[item]
                for date in dates:
                    date_str = date.strftime('%d/%m/%Y')
                    col = date_columns[date_str]
                    if job.dates.get(item) == date:
                        sheet.cell(row=row_index, column=col, value=job.scaffold[item])

            # Write deletions
            for item, deletions in job.deletions.items():
                if item not in items_written:
                    sheet.cell(row=row_num, column=1, value=item)
                    items_written[item] = row_num
                    row_num += 1
                row_index = items_written[item]
                for del_quantity, del_date in deletions:
                    date_str = del_date.strftime('%d/%m/%Y')
                    col = date_columns[date_str]
                    current_value = sheet.cell(row=row_index, column=col).value
                    if current_value:
                        sheet.cell(row=row_index, column=col, value=current_value - del_quantity)
                    else:
                        sheet.cell(row=row_index, column=col, value=-del_quantity)

        # Auto-size columns
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True

        workbook.save(self.file_name)
        print(f"Data successfully saved to {self.file_name}")
    
    def add_scaffold_data(self, new_data, date_str):
        try:
            workbook = load_workbook(self.file_name)
            sheet = workbook.active

            # Get the last column index
            max_col = sheet.max_column

            # Check if the date column already exists
            date_exists = False
            for col in range(2, max_col + 1):
                if sheet.cell(row=2, column=col).value == date_str:
                    date_exists = True
                    date_col = col
                    break

            # If date column doesn't exist, create a new one
            if not date_exists:
                date_col = max_col + 1
                sheet.cell(row=2, column=date_col).value = date_str

            # Create a dictionary of existing data
            existing_data = {}
            item_row = {}
            for row in range(3, sheet.max_row + 1):
                item = sheet.cell(row=row, column=1).value
                if item:
                    item = item.strip()
                    existing_data[item] = sheet.cell(row=row, column=date_col).value
                    item_row[item] = row

            # Update the existing items and add new items
            for item, qty in new_data.items():
                if item in existing_data:
                    sheet.cell(row=item_row[item], column=date_col).value = qty
                else:
                    new_row = sheet.max_row + 1
                    sheet.cell(row=new_row, column=1).value = item
                    sheet.cell(row=new_row, column=date_col).value = qty

            # Save the workbook
            workbook.save(self.file_name)
            print(f"New scaffold data successfully added for {date_str}")
        except Exception as e:
            print(f"An error occurred while adding scaffold data: {e}")